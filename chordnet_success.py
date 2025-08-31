#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import sys, subprocess, importlib.util, os, io, tempfile, random, re, threading, shutil, wave
from pathlib import Path

# ---------- Dependencies ----------
PIP_DEPS = ["openpyxl", "matplotlib", "numpy", "Pillow", "mido", "simpleaudio"]
SPEC_CHECK = {"Pillow":"PIL", "openpyxl":"openpyxl", "matplotlib":"matplotlib", "numpy":"numpy", "mido":"mido", "simpleaudio":"simpleaudio"}

def ensure_deps():
    if getattr(sys, "frozen", False):
        return
    missing = []
    for pip_name in PIP_DEPS:
        mod = SPEC_CHECK[pip_name]
        if importlib.util.find_spec(mod) is None:
            missing.append(pip_name)
    if missing:
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", *missing])
        except Exception as e:
            print("Automatic install failed, install manually:", " ".join(missing))

ensure_deps()

# ---------- Imports ----------
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from PIL import Image, ImageTk
from mido import MidiFile, MidiTrack, Message
import simpleaudio as sa

# ---------- Music theory ----------
note_to_semitone = {
    'C':0,  'B#':0, 'C#':1, 'Db':1, 'D':2, 'D#':3, 'Eb':3,
    'E':4,  'Fb':4, 'F':5,  'E#':5, 'F#':6, 'Gb':6, 'G':7,
    'G#':8, 'Ab':8, 'A':9,  'A#':10,'Bb':10,'B':11, 'Cb':11,
    'Ebb':2, 'Abb':6, 'Bbb':9, 'Fbb':3, 'Cbb':10, 'Gbb':5, 'Dbb':0
}
roman_regex = re.compile(r'^[b♭]?(?:I|II|III|IV|V|VI|VII)°?$', re.IGNORECASE)

def fallback_quality(numeral):
    n = (numeral or "").lower().replace('♭','b').replace('°','')
    return {'v':'dom7','ii':'min7','iii':'min7','vi':'min7'}.get(n, 'maj7')

def parse_chord_label(label):
    parts = (label or "").strip().split()
    if not parts:
        return None
    if roman_regex.match(parts[-1]):
        numeral = parts[-1]
        root_tokens = parts[:-1]
    else:
        numeral = ''
        root_tokens = parts
    chord_root = ''.join(root_tokens).replace(' ', '')
    if chord_root.endswith('°'):
        return chord_root[:-1], 'dim7'
    if chord_root.endswith('m'):
        return chord_root[:-1], 'min7'
    if (numeral or '').lower() == 'v':
        return chord_root, 'dom7'
    return chord_root, fallback_quality(numeral)

def build_7th_chord(root, quality):
    root_fixed = (root or "").replace('♯','#').replace('♭','b')
    r = note_to_semitone.get(root_fixed)
    if r is None:
        enharmonics = {'Ebb':'D','B#':'C','Cb':'B','Fb':'E','Abb':'G','Dbb':'C','Gbb':'F','Cbb':'Bb','Fbb':'Eb','Bbb':'A'}
        alt = enharmonics.get(root_fixed)
        if alt:
            r = note_to_semitone.get(alt)
    if r is None:
        return [], None
    intervals = {'maj7':[0,4,7,11],'min7':[0,3,7,10],'dom7':[0,4,7,10],'dim7':[0,3,6,9]}
    return [r+i for i in intervals.get(quality, [0,4,7,11])], r

CANONICAL_NOTE = {0:'C',1:'C#',2:'D',3:'Eb',4:'E',5:'F',6:'F#',7:'G',8:'Ab',9:'A',10:'Bb',11:'B'}
def canonical_note_name(root):
    root_fixed = (root or "").replace('♯','#').replace('♭','b')
    enh = {'Ebb':'D','Bbb':'A','Abb':'G','Dbb':'C','Gbb':'F','Cbb':'Bb','Fbb':'Eb','B#':'C','E#':'F','Cb':'B','Fb':'E'}
    if root_fixed in enh:
        return enh[root_fixed]
    if root_fixed in note_to_semitone:
        semi = note_to_semitone[root_fixed] % 12
        return CANONICAL_NOTE.get(semi, root)
    return root or "?"

def canonicalize_label(label):
    parsed = parse_chord_label(label)
    if not parsed:
        return label
    root, qual = parsed
    pretty = canonical_note_name(root)
    return "{} {}".format(pretty, qual)

# ---------- Drawing ----------
def draw_piano_image(notes, root_semi, width_px=420, height_px=110):
    base = 48
    whites, count, i = [], 0, base
    while count < 22:
        if i % 12 not in (1,3,6,8,10):
            whites.append(i); count += 1
        i += 1
    blacks = [n for n in range(base, base+36) if n % 12 in (1,3,6,8,10) and any(w < n < w+2 for w in whites)]
    fig, ax = plt.subplots(figsize=(max(1,width_px)/120, max(1,height_px)/120), dpi=120)
    onset = set([x % 12 for x in notes])
    for idx, m in enumerate(whites):
        on = (m % 12) in onset
        ax.add_patch(plt.Rectangle((idx,0),1,5, facecolor=('lightgreen' if on else 'white'), edgecolor='#222'))
        if root_semi is not None and (m % 12) == (root_semi % 12):
            ax.add_patch(plt.Circle((idx+0.5,4.1),0.25, color='#2563eb', zorder=11))
    for m in blacks:
        wi = next(i for i,w in enumerate(whites) if w > m) - 1
        x = wi + 0.65
        on = (m % 12) in onset
        ax.add_patch(plt.Rectangle((x,2.6),0.6,2.4, facecolor=('lightgreen' if on else '#111'),
                                   edgecolor='#222', zorder=10))
        if root_semi is not None and (m % 12) == (root_semi % 12):
            ax.add_patch(plt.Circle((x+0.3,4.2),0.18, color='#2563eb', zorder=12))
    ax.set_xlim(0, len(whites))
    ax.set_ylim(0, 5.3)
    ax.axis('off')
    plt.tight_layout(pad=0.1)
    buf = io.BytesIO()
    plt.savefig(buf, format='png')
    plt.close(fig)
    buf.seek(0)
    return Image.open(buf).convert("RGB")

# ---------- Audio ----------
def synthesize_chord_wav_bytes(notes, seconds=0.9, sr=44100):
    if not notes:
        return b"", sr
    base_midi = 48
    freqs = [440.0 * 2**(((base_midi + x) - 69)/12) for x in notes]
    t = np.linspace(0, seconds, int(sr*seconds), False)
    wave_f = np.zeros_like(t)
    for f in freqs:
        wave_f += np.sin(2 * np.pi * f * t)
    attack = int(0.02*sr)
    release = int(0.15*sr)
    env = np.ones_like(wave_f)
    env[:attack] = np.linspace(0,1,attack)
    env[-release:] = np.linspace(1,0,release)
    wave_f *= env
    peak = float(np.max(np.abs(wave_f))) if wave_f.size else 1.0
    wave_f = wave_f / (peak + 1e-9)
    wave_i16 = (wave_f * 32767).astype(np.int16)
    buf = io.BytesIO()
    with wave.open(buf, 'wb') as wf:
        wf.setnchannels(1)
        wf.setsampwidth(2)
        wf.setframerate(sr)
        wf.writeframes(wave_i16.tobytes())
    return buf.getvalue(), sr

def play_with_afplay_or_simpleaudio(notes):
    data, sr = synthesize_chord_wav_bytes(notes)
    if not data:
        return
    afplay = shutil.which("afplay")
    if afplay:
        try:
            tmp = tempfile.NamedTemporaryFile(prefix="chord_", suffix=".wav", delete=False)
            tmp.write(data); tmp.flush(); tmp.close()
            subprocess.Popen([afplay, tmp.name], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            threading.Timer(6.0, lambda p=tmp.name: (os.path.exists(p) and os.remove(p))).start()
            return
        except Exception as e:
            print("afplay failed, falling back to simpleaudio:", e)
    try:
        wave_obj = sa.WaveObject(data, 1, 2, sr)
        wave_obj.play()
    except Exception as e:
        print("simpleaudio failed:", e)

# ---------- MIDI ----------
def export_midi_for_chord(notes, out_path):
    if not notes:
        return
    mid = MidiFile()
    track = MidiTrack()
    mid.tracks.append(track)
    velocity = 90
    start = 0
    length = 480
    base_midi = 60
    for x in notes:
        pitch = base_midi + (x % 12)
        track.append(Message('note_on', note=pitch, velocity=velocity, time=start))
    for x in notes:
        pitch = base_midi + (x % 12)
        track.append(Message('note_off', note=pitch, velocity=64, time=length))
    mid.save(out_path)

# ---------- Excel parsing ----------
def load_workbook_structure(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    structured_data = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        key_data = {}
        current_base = None
        current_sub = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            base, sub = row[1], row[2]
            chords = [c for c in row[4:11] if c]
            if base:
                current_base = base
            if sub:
                current_sub = sub
            if current_base and current_sub and chords:
                key_data.setdefault(current_base, {}).setdefault(current_sub, []).append(chords)
        structured_data[sheet] = key_data
    return structured_data, wb.sheetnames

# ---------- App ----------
class ChordApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Chord Picker App")
        self.geometry("1280x720")
        self.minsize(980, 600)

        # ttk styling
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except Exception:
            pass
        self._style = style

        style.configure("TButton", padding=6, font=("Helvetica", 11))
        style.configure("Accent.TButton", background="#2563eb", foreground="#fff")
        style.map("Accent.TButton", background=[("active", "#1d4ed8")])
        style.configure("TLabel", font=("Helvetica", 11))
        style.configure("Header.TLabel", font=("Helvetica", 13, "bold"))
        style.configure("CardInner.TFrame", relief="flat")

        # State
        self.xlsx_path = None
        self.structured_data = {}
        self.sheet_names = []
        self.history = []            # undo (max 10)
        self.current_images = []
        self.current_chords = []     # list[str]

        # Variables
        self.key_var  = tk.StringVar(value="All Keys")
        self.base_var = tk.StringVar(value="All Base Modes")
        self.sub_var  = tk.StringVar(value="All Sub Modes")
        self.num_var  = tk.IntVar(value=8)

        # Controls (horizontal)
        top = ttk.Frame(self, padding=(10,10,10,6))
        top.pack(side=tk.TOP, fill=tk.X)

        row = ttk.Frame(top)
        row.pack(side=tk.TOP, fill=tk.X)

        self.load_btn = ttk.Button(row, text="Open Excel…", command=self.open_excel)
        self.load_btn.pack(side=tk.LEFT, padx=(0,8))

        self.key_cb  = ttk.Combobox(row, width=22, textvariable=self.key_var, state="readonly", values=["All Keys"])
        self.key_cb.pack(side=tk.LEFT, padx=6)
        self.base_cb = ttk.Combobox(row, width=22, textvariable=self.base_var, state="readonly", values=["All Base Modes"])
        self.base_cb.pack(side=tk.LEFT, padx=6)
        self.sub_cb  = ttk.Combobox(row, width=24, textvariable=self.sub_var, state="readonly", values=["All Sub Modes"])
        self.sub_cb.pack(side=tk.LEFT, padx=6)

        count_wrap = ttk.Frame(row); count_wrap.pack(side=tk.LEFT, padx=(12,4))
        ttk.Label(count_wrap, text="Chords").pack(side=tk.TOP, anchor="w")
        self.num_spin = ttk.Spinbox(count_wrap, from_=1, to=16, textvariable=self.num_var, width=6); self.num_spin.pack(side=tk.TOP, anchor="w")

        self.gen_btn = ttk.Button(row, text="Generate", command=self.generate_chords, style="Accent.TButton")
        self.gen_btn.pack(side=tk.LEFT, padx=(12,6))
        self.undo_btn = ttk.Button(row, text="Undo", command=self.undo, state="disabled")
        self.undo_btn.pack(side=tk.LEFT)

        self.dark_btn = ttk.Button(row, text="Dark Mode", command=self.toggle_theme)
        self.dark_btn.pack(side=tk.LEFT, padx=(12,0))

        # Scrollable results (grid container)
        self.canvas = tk.Canvas(self, borderwidth=0, highlightthickness=0)
        self.scroll_y = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.inner = ttk.Frame(self.canvas)
        self.inner.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.inner_id = self.canvas.create_window((0,0), window=self.inner, anchor="nw")
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.scroll_y.pack(side=tk.RIGHT, fill=tk.Y)

        # Bindings
        self.key_cb.bind("<<ComboboxSelected>>", self.on_key_change)
        self.base_cb.bind("<<ComboboxSelected>>", self.on_base_change)
        self.canvas.bind("<Configure>", self.on_canvas_resize)
        self.bind_all("<Key>", self.on_keypress)
        self._resize_job = None
        self._regen_mode = False

        # Theme
        self._dark = False
        self.apply_theme()

        # Autoload MMM7 and defaults
        self.try_autoload_xlsx_and_defaults()

    # ---------- Theme ----------
    def toggle_theme(self):
        self._dark = not self._dark
        self.apply_theme()
        # Update canvas bg and force a re-render for consistent look
        self.on_canvas_resize(type("e", (), {"width": self.canvas.winfo_width()})())
        if self.current_chords:
            self.render_results(self.current_chords)

    def apply_theme(self):
        style = self._style
        if self._dark:
            self.theme = {
                "bg":"#0f172a",        # slate-900
                "fg":"#e5e7eb",        # slate-200
                "card":"#111827",      # slate-800/900 mix
                "border":"#374151",    # slate-600 (softer than white)
            }
            self.dark_btn.configure(text="Light Mode")
        else:
            self.theme = {
                "bg":"#f8fafc",        # slate-50
                "fg":"#0f172a",        # slate-900
                "card":"#f9fafb",      # slate-50/100
                "border":"#e5e7eb",    # slate-200 (less pure white)
            }
            self.dark_btn.configure(text="Dark Mode")

        bg = self.theme["bg"]
        fg = self.theme["fg"]
        card = self.theme["card"]

        # Apply to root and common styles
        try: self.configure(bg=bg)
        except Exception: pass
        style.configure("TFrame", background=bg)
        style.configure("TLabel", background=bg, foreground=fg)
        style.configure("Header.TLabel", background=bg, foreground=fg)
        style.configure("CardInner.TFrame", background=card, relief="flat")

        # Canvas background to match
        try: self.canvas.configure(background=bg, highlightthickness=0, bd=0)
        except Exception: pass

    # ---------- History ----------
    def push_history(self):
        if self.current_chords:
            self.history.append(list(self.current_chords))
            if len(self.history) > 10:
                self.history.pop(0)
            self.undo_btn.config(state="normal")

    def undo(self):
        if not self.history:
            return
        self.current_chords = self.history.pop()
        self.undo_btn.config(state=("normal" if self.history else "disabled"))
        self.render_results(self.current_chords, from_undo=True)

    # ---------- Loading ----------
    def try_autoload_xlsx_and_defaults(self):
        cwd = Path.cwd()
        candidates = list(cwd.glob("MMM7*.xlsx")) + list(cwd.glob("MMM7*.XLSX"))
        if not candidates:
            candidates = list(cwd.glob("*.xlsx"))
        if candidates:
            self.load_excel(candidates[0])
            self.set_defaults_and_autogen()

    def open_excel(self):
        p = filedialog.askopenfilename(title="Select Excel file", filetypes=[("Excel files","*.xlsx")])
        if p:
            self.load_excel(Path(p))

    def load_excel(self, path: Path):
        try:
            data, sheets = load_workbook_structure(path)
        except Exception as e:
            messagebox.showerror("Error", "Failed to load workbook:\n{}".format(e))
            return
        self.xlsx_path = path
        self.structured_data = data
        self.sheet_names = list(sheets)
        self.key_cb["values"]  = ["All Keys"] + self.sheet_names
        self.key_var.set("All Keys")
        self.base_cb["values"] = ["All Base Modes"]
        self.base_var.set("All Base Modes")
        self.sub_cb["values"]  = ["All Sub Modes"]
        self.sub_var.set("All Sub Modes")
        self.clear_results()
        self.title("Chord Picker App — {}".format(path.name))

    def set_defaults_and_autogen(self):
        target_sheet = "C"
        if target_sheet in self.sheet_names:
            self.key_var.set(target_sheet)
            self.on_key_change()
            base = "C Ionian"
            sub  = "C Ionian Ionians"
            if base in self.base_cb["values"]:
                self.base_var.set(base); self.on_base_change()
            if sub in self.sub_cb["values"]:
                self.sub_var.set(sub)
            self.num_var.set(4)
            self.generate_chords()

    # ---------- Dropdown updates ----------
    def on_key_change(self, event=None):
        k = self.key_var.get()
        if k == "All Keys":
            self.base_cb["values"] = ["All Base Modes"]
            self.base_var.set("All Base Modes")
            self.sub_cb["values"]  = ["All Sub Modes"]
            self.sub_var.set("All Sub Modes")
        else:
            bases = sorted(self.structured_data.get(k, {}).keys())
            self.base_cb["values"] = ["All Base Modes"] + bases
            self.base_var.set("All Base Modes")
            self.sub_cb["values"]  = ["All Sub Modes"]
            self.sub_var.set("All Sub Modes")
        self.clear_results()

    def on_base_change(self, event=None):
        k = self.key_var.get()
        b = self.base_var.get()
        if k == "All Keys" or b == "All Base Modes":
            self.sub_cb["values"] = ["All Sub Modes"]
            self.sub_var.set("All Sub Modes")
        else:
            subs = sorted(self.structured_data[k].get(b, {}).keys())
            self.sub_cb["values"] = ["All Sub Modes"] + subs
            self.sub_var.set("All Sub Modes")
        self.clear_results()

    # ---------- Generate ----------
    def collect_pool(self, key, base, sub):
        all_chords = []
        keys = self.sheet_names if key == "All Keys" else [key]
        for k in keys:
            for b, subs in self.structured_data.get(k, {}).items():
                if base != "All Base Modes" and b != base:
                    continue
                for s, rows in subs.items():
                    if sub != "All Sub Modes" and s != sub:
                        continue
                    for row in rows:
                        all_chords.extend([c for c in row if c])
        return all_chords

    def generate_chords(self):
        if not self.structured_data:
            messagebox.showinfo("Load Excel", "Open your .xlsx first.")
            return
        key, base, sub = self.key_var.get(), self.base_var.get(), self.sub_var.get()
        num = max(1, int(self.num_var.get()))
        pool = self.collect_pool(key, base, sub)
        if not pool:
            self.clear_results()
            ttk.Label(self.inner, text="No chords found for this filter.", foreground="#a00").grid(row=0, column=0, sticky="w", padx=12, pady=10)
            return
        self.push_history()
        sample = random.sample(pool, min(num, len(pool)))
        self.current_chords = sample
        self.render_results(sample)

    def regen_single(self, index):
        key, base, sub = self.key_var.get(), self.base_var.get(), self.sub_var.get()
        pool = self.collect_pool(key, base, sub)
        if not pool:
            return
        self.push_history()
        existing = set(self.current_chords)
        options = [c for c in pool if c not in existing] or pool
        if 0 <= index < len(self.current_chords):
            self.current_chords[index] = random.choice(options)
            self.render_results(self.current_chords)

    # ---------- Rendering ----------
    def clear_results(self):
        for w in list(self.inner.children.values()):
            w.destroy()
        self.current_images.clear()

    def render_results(self, chords, from_undo=False):
        self.clear_results()
        try:
            avail = getattr(self, "_canvas_width", self.canvas.winfo_width())
            if not avail or avail <= 1:
                avail = max(0, self.winfo_width() - 30)
        except Exception:
            avail = 1200
        total_padding = 40
        col_w = max(360, int((avail - total_padding) / 2))
        self.inner.grid_columnconfigure(0, weight=1, uniform="col")
        self.inner.grid_columnconfigure(1, weight=1, uniform="col")

        border_color = self.theme["border"]
        for idx, chord in enumerate(chords):
            parsed = parse_chord_label(chord)
            if not parsed:
                continue
            root, qual = parsed
            notes, root_semi = build_7th_chord(root, qual)
            r, c = idx // 2, idx % 2

            # Outer border frame (tk.Frame so we can color the "border")
            outer = tk.Frame(self.inner, bg=border_color, highlightthickness=0, bd=0)
            outer.grid(row=r, column=c, sticky="nwe", padx=10, pady=8)

            # Inner card
            card = ttk.Frame(outer, padding=(12,10), style="CardInner.TFrame")
            card.pack(fill=tk.BOTH, expand=True, padx=1, pady=1)

            head = ttk.Frame(card); head.pack(fill=tk.X)
            ttk.Label(head, text="#{}".format(idx+1), style="Header.TLabel").pack(side=tk.LEFT, padx=(0,8))
            ttk.Label(head, text=canonicalize_label(chord), style="Header.TLabel").pack(side=tk.LEFT)

            controls_est = 180
            img_w = max(260, col_w - controls_est - 40)
            img_h = int(img_w * (110/420))
            img = draw_piano_image(notes, root_semi, width_px=img_w, height_px=img_h)
            imgtk = ImageTk.PhotoImage(img)
            self.current_images.append(imgtk)
            ttk.Label(card, image=imgtk).pack(side=tk.LEFT, padx=(0,12))

            controls = ttk.Frame(card); controls.pack(side=tk.LEFT, fill=tk.Y)
            ttk.Button(controls, text="Play",
                       command=lambda n=notes: threading.Thread(target=play_with_afplay_or_simpleaudio, args=(n,), daemon=True).start()
                       ).pack(fill=tk.X, pady=(0,6))
            ttk.Button(controls, text="Regen", command=lambda i=idx: self.regen_single(i)).pack(fill=tk.X, pady=(0,6))
            ttk.Button(controls, text="Export MIDI",
                       command=lambda n=notes, name=chord: self.export_midi(n, name)
                       ).pack(fill=tk.X)

        self.undo_btn.config(state=("normal" if self.history else "disabled"))

    # ---------- Resize & hotkeys ----------
    def on_canvas_resize(self, event):
        try:
            self.canvas.itemconfigure(self.inner_id, width=event.width)
            self._canvas_width = event.width
        except Exception:
            self._canvas_width = event.width
        if getattr(self, "_resize_job", None):
            self.after_cancel(self._resize_job)
        self._resize_job = self.after(80, self._apply_resize)

    def _apply_resize(self):
        self._resize_job = None
        if self.current_chords:
            self.render_results(self.current_chords)

    def on_keypress(self, event):
        ch = event.char
        if not ch:
            return
        ch = ch.lower()
        if ch == 'u':
            self.undo(); return
        if ch == 'g':
            self.generate_chords(); return
        if ch == 'r':
            self._regen_mode = True; return
        if ch.isdigit():
            idx = int(ch) - 1
            if 0 <= idx < min(8, len(self.current_chords)):
                if getattr(self, "_regen_mode", False):
                    self._regen_mode = False
                    self.regen_single(idx)
                else:
                    self.play_index(idx)

    def play_index(self, index):
        try:
            chord = self.current_chords[index]
        except IndexError:
            return
        parsed = parse_chord_label(chord)
        if not parsed:
            return
        root, qual = parsed
        notes, _ = build_7th_chord(root, qual)
        threading.Thread(target=play_with_afplay_or_simpleaudio, args=(notes,), daemon=True).start()

    # ---------- Actions ----------
    def export_midi(self, notes, chord_name):
        if not self.xlsx_path:
            return
        safe = re.sub(r'[^A-Za-z0-9_\-]+', '_', chord_name).strip("_")
        out = self.xlsx_path.with_suffix("").parent / "{}.mid".format(safe)
        try:
            export_midi_for_chord(notes, out)
            messagebox.showinfo("MIDI Exported", "Saved: {}".format(out))
        except Exception as e:
            messagebox.showerror("Error", "Failed to export MIDI:\n{}".format(e))

def main():
    app = ChordApp()
    app.mainloop()

if __name__ == "__main__":
    main()
